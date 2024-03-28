import pdfplumber
from openpyxl import Workbook

# 读取PDF文件并将其转换为DataFrame
pdfFile = "/Users/echo/Downloads/反恐验厂不符合项整改报告_excel.pdf"
outPutFile = "output.xlsx"

def convertPdfToExcel(pdfFile, outPutFile):
    # 打开PDF文件
    pdf_file = open(pdfFile, 'rb')
    pdf_reader = pdfplumber.open(pdf_file)

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    row_num = 1
    # 逐页提取PDF内容并写入Excel
    for page in pdf_reader.pages:
        text = page.extract_text()
        lines = text.split("\n")
        # print(text)
        # exit(0)

        for line in lines:
            space_count = line.count(" ")
            if space_count > 8:
                ws.cell(row=row_num, column=1, value=line)
            else:
                columnVals = line.split("  ")
                for colu, val in enumerate(columnVals):
                    ws.cell(row=row_num, column=colu+1, value=val)
            row_num = row_num + 1

    # 保存Excel文件
    wb.save(outPutFile)
